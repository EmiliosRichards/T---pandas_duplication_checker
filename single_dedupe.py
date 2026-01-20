import pandas as pd
import os
import argparse
from openpyxl.utils import get_column_letter
from typing import Optional, Tuple

# --- Configuration ---
INPUT_FILE_PATH = r"data\blist_003_AS700_20250625.xlsx"  # IMPORTANT: Change this to your input file path
DEDUPLICATION_COLUMN = "Company Name"  # Column to use for identifying duplicates
OUTPUT_DIR = "single_output"
LOG_DIR = "single_logs"

def _sniff_csv_separator(input_file_path: str, default: str = ';') -> str:
    """
    Best-effort delimiter detection (prefers ';' for EU-style CSVs).
    """
    try:
        with open(input_file_path, "r", encoding="utf-8-sig", errors="ignore") as f:
            header = f.readline()
        if header.count(';') >= header.count(',') and header.count(';') > 0:
            return ';'
        if header.count(',') > 0:
            return ','
    except Exception:
        pass
    return default

def _load_dataframe(input_file_path: str) -> Tuple[pd.DataFrame, Optional[str]]:
    """
    Loads CSV or Excel into a DataFrame.

    Returns:
      (df, csv_sep) where csv_sep is only set for CSV inputs.
    """
    _, ext = os.path.splitext(input_file_path)
    ext = ext.lower()

    if ext in (".xlsx", ".xls", ".xlsm"):
        return pd.read_excel(input_file_path, dtype=str), None

    if ext in (".csv", ".txt"):
        sep = _sniff_csv_separator(input_file_path)
        return pd.read_csv(input_file_path, sep=sep, dtype=str, encoding="utf-8-sig"), sep

    raise ValueError(f"Unsupported input file type: '{ext}'. Use .csv or Excel (.xlsx/.xls/.xlsm).")

def format_excel_sheet(writer, df, sheet_name="Sheet1"):
    """Applies formatting to the Excel sheet."""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    # 1. Format 'Number' column as text, if it exists
    if "Number" in df.columns:
        try:
            number_col_idx = df.columns.get_loc("Number")
            number_col_letter = get_column_letter(number_col_idx + 1)  # openpyxl columns are 1-based

            for row_idx in range(2, len(df) + 2):  # Skip header, Excel rows are 1-based
                cell = worksheet[f"{number_col_letter}{row_idx}"]
                cell.number_format = "@"
        except KeyError:
            print(f"Warning: Column 'Number' not found for text formatting.")
        except Exception as e:
            print(f"Error formatting 'Number' column: {e}")


    # 2. Auto-adjust column widths
    for col_idx, column_name in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        # Check column header length
        if column_name is not None:
            max_length = max(max_length, len(str(column_name)))

        # Check cell content length
        # Iterate through the column cells in the worksheet
        # worksheet.columns is an iterator of tuples of cells for each column
        # We need to access the correct column by its letter or index
        # For simplicity, we iterate through DataFrame column values
        if not df[column_name].empty:
             max_length = max(
                max_length,
                df[column_name].astype(str).map(len).max()
            )
        
        adjusted_width = max_length + 2  # Adding a little padding
        worksheet.column_dimensions[column_letter].width = adjusted_width

def _default_output_paths(input_file_path: str, output_dir: str, log_dir: str) -> Tuple[str, str]:
    base_name, ext = os.path.splitext(os.path.basename(input_file_path))
    cleaned_file_name = f"{base_name}_deduped{ext}"
    removed_log_file_name = f"{base_name}_deduped_removed_log{ext}"
    return (
        os.path.join(output_dir, cleaned_file_name),
        os.path.join(log_dir, removed_log_file_name),
    )

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Deduplicate a CSV or Excel file based on a single column, keeping the first occurrence. "
            "Writes a cleaned output file plus a *_removed_log file containing the removed duplicate rows."
        )
    )
    parser.add_argument("-i", "--input", default=INPUT_FILE_PATH, help="Input file path (.csv/.xlsx/.xls/.xlsm).")
    parser.add_argument(
        "-c",
        "--dedupe-column",
        default=DEDUPLICATION_COLUMN,
        help="Column name to use for identifying duplicates.",
    )
    parser.add_argument(
        "--output-dir",
        default=OUTPUT_DIR,
        help="Directory for the cleaned output (used only if --output is not provided).",
    )
    parser.add_argument(
        "--log-dir",
        default=LOG_DIR,
        help="Directory for the removed-rows log (used only if --removed-log is not provided).",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Optional explicit output file path for the deduped file. Extension controls output format.",
    )
    parser.add_argument(
        "--removed-log",
        default=None,
        help="Optional explicit output file path for the removed duplicates log.",
    )

    args = parser.parse_args()

    input_path = args.input
    dedupe_col = args.dedupe_column

    output_dir = args.output_dir
    log_dir = args.log_dir
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(log_dir, exist_ok=True)

    cleaned_output_path, removed_log_path = _default_output_paths(input_path, output_dir, log_dir)
    if args.output:
        cleaned_output_path = args.output
    if args.removed_log:
        removed_log_path = args.removed_log

    print(f"Starting deduplication process for: {input_path}")

    # --- Load the input file ---
    try:
        df_original, csv_sep = _load_dataframe(input_path)
        print(f"Successfully loaded {input_path}. Original rows: {len(df_original)}")
    except FileNotFoundError:
        print(f"Error: Input file not found at {input_path}")
        return
    except ValueError as e:
        print(f"Error: {e}")
        return
    except Exception as e:
        print(f"Error loading input file: {e}")
        return

    if dedupe_col not in df_original.columns:
        print(f"Error: Deduplication column '{dedupe_col}' not found in the input file.")
        print(f"Available columns are: {df_original.columns.tolist()}")
        return

    # --- Identify and separate duplicates ---
    # Keep the first occurrence, mark others as duplicates
    duplicates_mask = df_original.duplicated(subset=[dedupe_col], keep='first')
    
    df_cleaned = df_original[~duplicates_mask]
    df_removed = df_original[duplicates_mask]

    print(f"Deduplication based on column: '{dedupe_col}'")
    print(f"Number of rows removed as duplicates: {len(df_removed)}")
    print(f"Number of rows remaining after deduplication: {len(df_cleaned)}")

    # --- Save the cleaned data ---
    try:
        _, out_ext = os.path.splitext(cleaned_output_path)
        out_ext = out_ext.lower()

        if out_ext == ".csv":
            sep_to_use = csv_sep or ';'
            df_cleaned.to_csv(cleaned_output_path, index=False, sep=sep_to_use, encoding="utf-8-sig")
        else:
            with pd.ExcelWriter(cleaned_output_path, engine="openpyxl") as writer:
                df_cleaned.to_excel(writer, index=False, sheet_name="Sheet1")
                format_excel_sheet(writer, df_cleaned, sheet_name="Sheet1")

        print(f"Cleaned data saved to: {cleaned_output_path}")
    except Exception as e:
        print(f"Error saving cleaned data: {e}")

    # --- Save the removed duplicates (log) ---
    if not df_removed.empty:
        try:
            _, log_ext = os.path.splitext(removed_log_path)
            log_ext = log_ext.lower()

            if log_ext == ".csv":
                sep_to_use = csv_sep or ';'
                df_removed.to_csv(removed_log_path, index=False, sep=sep_to_use, encoding="utf-8-sig")
            else:
                with pd.ExcelWriter(removed_log_path, engine="openpyxl") as writer:
                    df_removed.to_excel(writer, index=False, sheet_name="RemovedDuplicates")
                    format_excel_sheet(writer, df_removed, sheet_name="RemovedDuplicates")

            print(f"Removed duplicates logged to: {removed_log_path}")
        except Exception as e:
            print(f"Error saving removed duplicates log: {e}")
    else:
        print("No duplicates found to log.")

    print("\n--- Summary ---")
    print(f"Original rows: {len(df_original)}")
    print(f"Rows removed: {len(df_removed)}")
    print(f"Final cleaned rows: {len(df_cleaned)}")
    print("--- Process Complete ---")

if __name__ == "__main__":
    main()