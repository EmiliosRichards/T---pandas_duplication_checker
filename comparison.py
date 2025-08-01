import pandas as pd
import os
from datetime import datetime # Keep for potential future use or more detailed logging if needed
from openpyxl.utils import get_column_letter

# --- Configuration ---
# IMPORTANT: Change these to your input file paths
FILE_TO_BE_MODIFIED_PATH = r"single_output\Adressen B 25.06.25_deduped_deduped.xlsx"
FILE_TO_COMPARE_AGAINST_PATH = r"data\blist_003_AS250_rbotf_20250626.xlsx"

COMPARISON_COLUMN = "URL"  # Column to use for matching
OUTPUT_DIR = "comparison_output"
LOG_DIR = "comparison_logs"

# --- Ensure output and log directories exist ---
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(LOG_DIR, exist_ok=True)

def format_excel_sheet(writer, df, sheet_name="Sheet1"):
    """Applies formatting to the Excel sheet."""
    # workbook = writer.book # Not directly used here, but good to know it's available
    worksheet = writer.sheets[sheet_name]

    # 1. Format 'Number' column as text, if it exists
    if "Number" in df.columns:
        try:
            number_col_idx = df.columns.get_loc("Number")
            number_col_letter = get_column_letter(number_col_idx + 1)  # openpyxl columns are 1-based

            for row_idx in range(2, len(df) + 2):  # Skip header, Excel rows are 1-based
                cell = worksheet[f"{number_col_letter}{row_idx}"]
                cell.number_format = "@"
        except KeyError:
            print(f"Warning: Column 'Number' not found for text formatting in sheet '{sheet_name}'.")
        except Exception as e:
            print(f"Error formatting 'Number' column in sheet '{sheet_name}': {e}")

    # 2. Auto-adjust column widths
    for col_idx, column_name in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        if column_name is not None:
            max_length = max(max_length, len(str(column_name)))

        if not df[column_name].empty:
             max_length = max(
                max_length,
                df[column_name].astype(str).map(len).max()
            )
        
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def main():
    print(f"Starting comparison process.")
    print(f"File to be modified: {FILE_TO_BE_MODIFIED_PATH}")
    print(f"File to compare against: {FILE_TO_COMPARE_AGAINST_PATH}")

    # --- Load the Excel files ---
    try:
        df_getting_changed = pd.read_excel(FILE_TO_BE_MODIFIED_PATH, dtype=str)
        print(f"Successfully loaded '{FILE_TO_BE_MODIFIED_PATH}'. Original rows: {len(df_getting_changed)}")
    except FileNotFoundError:
        print(f"Error: Input file not found at {FILE_TO_BE_MODIFIED_PATH}")
        return
    except Exception as e:
        print(f"Error loading Excel file '{FILE_TO_BE_MODIFIED_PATH}': {e}")
        return

    try:
        df_compared_to = pd.read_excel(FILE_TO_COMPARE_AGAINST_PATH, dtype=str)
        print(f"Successfully loaded '{FILE_TO_COMPARE_AGAINST_PATH}'. Rows: {len(df_compared_to)}")
    except FileNotFoundError:
        print(f"Error: Input file not found at {FILE_TO_COMPARE_AGAINST_PATH}")
        return
    except Exception as e:
        print(f"Error loading Excel file '{FILE_TO_COMPARE_AGAINST_PATH}': {e}")
        return

    # --- Ensure comparison column exists in both DataFrames ---
    if COMPARISON_COLUMN not in df_getting_changed.columns:
        print(f"Error: Comparison column '{COMPARISON_COLUMN}' not found in '{FILE_TO_BE_MODIFIED_PATH}'.")
        print(f"Available columns: {df_getting_changed.columns.tolist()}")
        return
    if COMPARISON_COLUMN not in df_compared_to.columns:
        print(f"Error: Comparison column '{COMPARISON_COLUMN}' not found in '{FILE_TO_COMPARE_AGAINST_PATH}'.")
        print(f"Available columns: {df_compared_to.columns.tolist()}")
        return

    # --- Optional: Drop internal duplicates within each file by the comparison column before comparison ---
    # This prevents issues if a URL appears multiple times within the same file
    df_getting_changed_orig_len = len(df_getting_changed)
    df_compared_to_orig_len = len(df_compared_to)

    df_getting_changed = df_getting_changed.drop_duplicates(subset=[COMPARISON_COLUMN], keep='first')
    df_compared_to = df_compared_to.drop_duplicates(subset=[COMPARISON_COLUMN], keep='first')

    if len(df_getting_changed) < df_getting_changed_orig_len:
        print(f"Note: Removed {df_getting_changed_orig_len - len(df_getting_changed)} internal duplicates from '{os.path.basename(FILE_TO_BE_MODIFIED_PATH)}' based on '{COMPARISON_COLUMN}'.")
    if len(df_compared_to) < df_compared_to_orig_len:
        print(f"Note: Removed {df_compared_to_orig_len - len(df_compared_to)} internal duplicates from '{os.path.basename(FILE_TO_COMPARE_AGAINST_PATH)}' based on '{COMPARISON_COLUMN}'.")


    # --- Step 1: Identify rows in df_getting_changed that have matching URLs in df_compared_to ---
    matching_rows_mask = df_getting_changed[COMPARISON_COLUMN].isin(df_compared_to[COMPARISON_COLUMN])
    df_matching_removed = df_getting_changed[matching_rows_mask]

    # --- Step 2: Remove those rows from df_getting_changed ---
    df_result_cleaned = df_getting_changed[~matching_rows_mask]

    print(f"\nComparison based on column: '{COMPARISON_COLUMN}'")
    print(f"Number of rows found in '{os.path.basename(FILE_TO_BE_MODIFIED_PATH)}' that match '{os.path.basename(FILE_TO_COMPARE_AGAINST_PATH)}': {len(df_matching_removed)}")
    print(f"Number of rows remaining in '{os.path.basename(FILE_TO_BE_MODIFIED_PATH)}' after removal: {len(df_result_cleaned)}")

    # --- Define output file paths based on the first input file name ---
    base_name_modified, ext_modified = os.path.splitext(os.path.basename(FILE_TO_BE_MODIFIED_PATH))
    
    cleaned_output_filename = f"{base_name_modified}_comparison_removed{ext_modified}"
    CLEANED_OUTPUT_PATH = os.path.join(OUTPUT_DIR, cleaned_output_filename)
    
    matching_log_filename = f"{base_name_modified}_comparison_matches_log{ext_modified}"
    MATCHING_LOG_PATH = os.path.join(LOG_DIR, matching_log_filename)

    # --- Save the cleaned (result) data ---
    try:
        with pd.ExcelWriter(CLEANED_OUTPUT_PATH, engine="openpyxl") as writer:
            df_result_cleaned.to_excel(writer, index=False, sheet_name="CleanedData")
            format_excel_sheet(writer, df_result_cleaned, sheet_name="CleanedData")
        print(f"✅ Cleaned data (after removing matches) saved to: {CLEANED_OUTPUT_PATH}")
    except Exception as e:
        print(f"Error saving cleaned data: {e}")

    # --- Save the matching/removed rows (log) ---
    if not df_matching_removed.empty:
        try:
            with pd.ExcelWriter(MATCHING_LOG_PATH, engine="openpyxl") as writer:
                df_matching_removed.to_excel(writer, index=False, sheet_name="MatchingRemovedRows")
                format_excel_sheet(writer, df_matching_removed, sheet_name="MatchingRemovedRows")
            print(f"ℹ️ Rows that matched and were removed logged to: {MATCHING_LOG_PATH}")
        except Exception as e:
            print(f"Error saving matching/removed rows log: {e}")
    else:
        print(f"ℹ️ No matching rows found to log from '{os.path.basename(FILE_TO_BE_MODIFIED_PATH)}'.")

    print("\n--- Summary ---")
    print(f"Initial rows in '{os.path.basename(FILE_TO_BE_MODIFIED_PATH)}' (after internal dedupe): {len(df_getting_changed)}")
    print(f"Rows removed due to match with '{os.path.basename(FILE_TO_COMPARE_AGAINST_PATH)}': {len(df_matching_removed)}")
    print(f"Final rows in cleaned output: {len(df_result_cleaned)}")
    print("--- Process Complete ---")

if __name__ == "__main__":
    main()